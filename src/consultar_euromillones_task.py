from datetime import datetime, timedelta
from airflow import DAG
from airflow.operators.python import PythonOperator
from bs4 import BeautifulSoup
import requests
import openpyxl

def saludar():
    print("Holaaa Airflow")

def extraer_numeros_premiados():
    fecha = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    url = "https://www.euromillones.com.es/"
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        meta_description = soup.find('meta', attrs={'name': 'description'})['content']

        numeros_str = meta_description.split('Números: ')[1].split(' Estrellas: ')[0]
        numeros = [int(num) for num in numeros_str.split(',')]
        estrellas_str = meta_description.split('Estrellas: ')[1]
        estrellas = [int(est) for est in estrellas_str.split(',')]
        workbook = openpyxl.load_workbook("/home/olga/euromillones_premiados.xlsx")
        sheet = workbook.active
        sheet.append(["Fecha", "Números Premiados", "Estrella 1", "Estrella 2"])

        fechas_existentes = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
        if fecha not in fechas_existentes:
            if len(estrellas) >= 2:
                sheet.append([fecha, ', '.join(map(str, numeros)), estrellas[0], estrellas[1]])
            else:
                sheet.append([fecha, ', '.join(map(str, numeros)), None, None])  
            workbook.save("euromillones_premiados.xlsx")
            print(f"Datos guardados en euromillones_premiados.xlsx con la última actualización que consta en la web {fecha}")
        else:
            print(f"Los datos ya estaban guardados anteriormente con fecha de {fecha}. No han habido cambios nuevos para guardar.")
    else:
        print("No se pudo acceder a la página. Código de estado:", response.status_code)


if __name__ == "__main__":
    extraer_numeros_premiados()